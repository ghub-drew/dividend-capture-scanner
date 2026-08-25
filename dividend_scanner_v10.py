"""
Dividend Capture Scanner  (v10)
Pulls upcoming ex-dividend dates from Yahoo Finance and ranks the best
dividend-capture candidates.

What it looks at:
  - Dividend yield        : ranked directly by percentage (higher is better).
  - Rebound time          : average AND worst-case trading days for the price
                            to climb back to its pre-ex-dividend level. Events
                            that never recovered show as ">60" and are counted
                            against the stock, not hidden.
  - Recovery rate         : how often it recovered within the test window.
  - Best entry timing     : back-tests buying 1 to 20 trading days before the
                            ex-date to find each stock's historical sweet spot.
  - Drop ratio            : how much the price actually falls on the ex-date
                            compared with the dividend paid. Below 1.0 means
                            the stock gives up less than the dividend.
  - Price trend           : up / down / mixed, from moving averages on
                            dividend-adjusted (total-return) prices, so big
                            payers aren't unfairly flagged as falling.
  - Sector                : each name's sector, so whole sectors can be ruled
                            out (see EXCLUDED_SECTORS) and so the plan can warn
                            when several trades pile into the same one.
  - Liquidity             : flags thinly-traded names that are harder to exit.
  - Round-lot economics   : dividend dollars and capital required per 100 shares.
  - After-tax gain        : dividend-capture trades are short-term, taxed as
                            ordinary income. The exact rate depends on
                            deductions and the year's tax rules, so every
                            after-tax figure is shown at two rates side by
                            side (see TAX_RATE_HIGH / TAX_RATE_LOW below).
  - Annualised efficiency : models continuous capital rotation - hold for the
                            rebound, sell at even-or-better, redeploy into the
                            next dividend. Failed recoveries are charged at the
                            full 60-day wait, so unreliable stocks score lower.
  - Rotation plan         : a chronological trade calendar for one or more
                            independent pots of capital (see CAPITAL and
                            N_SLICES), with estimated net dollars per trade
                            after tax and any financing fee. The plan knows
                            what is already held (see HOLDINGS) so it never
                            proposes a stock the account already owns and never
                            schedules a pot whose money is still tied up.
  - Preferred list        : a quality whitelist (see FAVORED_TICKERS) of names
                            judged safe to keep holding if a rebound takes far
                            longer than expected. They are flagged in the
                            output and win close calls in the Rotation Plan.
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from math import ceil
import warnings
warnings.filterwarnings("ignore")

# ─── S&P 500 tickers (fetched from Wikipedia) ───
def get_sp500_tickers():
    import requests
    from io import StringIO
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; dividend-scanner/1.0)"}
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    tables = pd.read_html(StringIO(resp.text))
    symbols = tables[0]["Symbol"].tolist()
    # Yahoo uses dashes where Wikipedia uses dots (BRK.B -> BRK-B)
    return [s.replace(".", "-") for s in symbols]

# ─── High-yield income names (mostly outside the S&P 500) ───
HIGH_YIELD_TICKERS = [
    # Mortgage REITs (often 9-15% yield)
    "AGNC", "NLY", "STWD", "ABR", "RITM", "DX", "CIM", "TWO", "PMT", "ARI",
    "EFC", "MFA", "IVR", "ORC",
    # Business Development Companies (BDCs, often 8-12%)
    "ARCC", "MAIN", "FSK", "OBDC", "BXSL", "PSEC", "HTGC", "GBDC", "TSLX",
    "CSWC", "GAIN", "PFLT", "BBDC", "TCPC",
    # REITs (5-9% yield)
    "O", "WPC", "STAG", "NNN", "EPR", "GLPI", "OHI", "LTC", "SBRA", "ADC",
    "IRM", "VICI", "KRG", "GNL",
    # Energy infrastructure (6-8%)
    "KMI", "OKE", "WES", "ET", "EPD", "PAA", "MPLX", "ENB",
    # Other high-yield blue chips
    "BTI", "VZ", "T", "MO", "PFE",
]

# Closed-end funds (CEFs) - commonly high yield (7-11%).
# Note: CLM and CRF were removed; their headline yields are largely a return
# of the investor's own capital, which makes them misleading for capture.
CEF_TICKERS = [
    "PDI", "PTY", "PDO", "UTF", "UTG", "USA", "ETV", "ETY", "BST", "DSL",
    "RA", "GOF", "ECC", "OXLC", "EVT", "BME", "HQH", "GAB",
    "PCN", "PFL", "BCAT",
]

# Preferred-stock exposure via clean ETF tickers (individual preferred symbols
# are unreliable on Yahoo; these are a dependable proxy for the asset class).
PREFERRED_TICKERS = ["PFF", "PFFA", "PFFD", "PGX", "PGF", "FPE", "PFXF"]

# ─── Preferred names (quality whitelist) ───
# Companies and funds judged safe to keep holding for an extended period if a
# rebound takes far longer than expected (the "would you be comfortable being
# stuck with it in a crash" test). They get a Pref flag in the output and win
# close calls in the Rotation Plan (see PREFERRED_EDGE below). The displayed
# Score is NOT changed - the preference only applies when picking trades.
FAVORED_TICKERS = [
    "NLY", "HBAN", "O", "ARCC", "BTI", "VZ", "MO", "HRL", "ARES", "DTE",
    "TROW", "CSWC", "ETV", "CPB", "EVT", "DSL", "GOF", "ETY",
]
PREFERRED_EDGE = 0.10   # a preferred name wins the plan pick unless a
                        # non-preferred candidate scores >10% higher

def get_universe():
    """S&P 500 + high-yield names + CEFs + preferreds, de-duplicated."""
    sp500 = get_sp500_tickers()
    combined = sp500 + HIGH_YIELD_TICKERS + CEF_TICKERS + PREFERRED_TICKERS
    return list(dict.fromkeys(combined))  # preserves order, dedupes

# ─── Filter config (adjust to taste) ───
MIN_ANNUAL_YIELD     = 0.03      # 3% floor
MIN_AVG_VOLUME       = 100_000   # lowered so CEFs/preferreds appear; thin ones get flagged
MIN_STOCK_PRICE      = 5.0       # dollars
MAX_STOCK_PRICE      = 500.0     # dollars
EX_DATE_WINDOW_DAYS  = 30        # how far ahead to look (wide enough to plan rotations)

# ─── Sector config ───
# Sectors listed here are dropped from the results entirely. The scan reports
# how many names were removed, so an excluded sector is never silently missing.
#
# Honest caveat, worth keeping in mind before adding to this list: excluding a
# sector removes the names, not the risk behind them. Interest rates move
# Real Estate, Utilities, the BDCs (which Yahoo files under Financial Services)
# and bond funds alike, so ruling out one of them narrows the candidate list
# without removing the exposure that comes with high yields.
EXCLUDED_SECTORS = ["Real Estate"]

# Funds and ETFs have no sector on Yahoo, so they are labelled by type here.
FUND_SECTORS = {t: "Fund - CEF" for t in CEF_TICKERS}
FUND_SECTORS.update({t: "Fund - Preferred" for t in PREFERRED_TICKERS})

# ─── Economics / tax config ───
# Dividend-capture trades are too short for the lower long-term capital-gains
# rate, so they are taxed as ordinary income. The exact rate depends on
# deductions and the year's tax rules, so every after-tax figure is shown at
# TWO rates side by side. The scoring math uses the higher (conservative) one.
TAX_RATE_HIGH   = 0.35   # conservative estimate of the effective rate
TAX_RATE_LOW    = 0.25   # optimistic estimate of the effective rate
ROUND_LOT       = 100    # shares - round lots are far easier to sell than odd lots
BUY_BUFFER_DAYS = 2      # days capital is tied up before the ex-date (rotation calc)
TRADING_DAYS    = 252    # trading days per year

# Column labels follow the rates above, so changing a rate renames the columns.
COL_AT_HIGH  = f"AfterTax{int(TAX_RATE_HIGH * 100)}/100"
COL_AT_LOW   = f"AfterTax{int(TAX_RATE_LOW * 100)}/100"
COL_NET_HIGH = f"Net {int(TAX_RATE_HIGH * 100)}% $"
COL_NET_LOW  = f"Net {int(TAX_RATE_LOW * 100)}% $"
COL_RUN_HIGH = f"Running {int(TAX_RATE_HIGH * 100)}% $"
COL_RUN_LOW  = f"Running {int(TAX_RATE_LOW * 100)}% $"

# ─── Rotation-plan config ───
# CAPITAL is the size of ONE slice, not the whole account. A slice is a pot of
# money that holds one position at a time and rotates independently of the
# others, so a position that gets stuck below its entry price freezes its own
# slice and leaves the rest still trading.
CAPITAL         = 50_000  # dollars per slice
N_SLICES        = 4       # independent pots (total committed = CAPITAL x N_SLICES)
FINANCE_FEE     = 0.00    # cut taken by a trade-financing company, if any
                          # (e.g. 0.10 = they keep 10% of the gross dividend)
SETTLEMENT_DAYS = 1       # trading days for sale proceeds to settle (T+1)

# ─── Current holdings ───
# What the account is holding RIGHT NOW. The plan reads this so it cannot
# propose a stock already owned and cannot schedule a slice whose money is
# still tied up. Keep it current: an out-of-date list produces a plan that
# looks fine and is not fundable.
#   slice  : which pot the position sits in (1 to N_SLICES)
#   free   : the date the cash is expected back (YYYY-MM-DD), or None if the
#            position is stuck below its entry price with no sell date, which
#            blocks that slice until it is sold.
HOLDINGS = [
    {"slice": 1, "ticker": "O",    "shares": 700,  "entry": 65.74, "free": None},
    {"slice": 3, "ticker": "KHC",  "shares": 1900, "entry": 25.44, "free": "2026-09-08"},
    {"slice": 4, "ticker": "AMCR", "shares": 1000, "entry": 45.40, "free": "2026-09-08"},
]

# ─── Liquidity rating thresholds (avg shares/day) ───
THIN_VOLUME = 500_000
HIGH_VOLUME = 2_000_000

# ─── Rebound-metric config ───
REBOUND_LOOKBACK_EVENTS = 8     # measure the last N ex-dividend events
REBOUND_MAX_DAYS        = 60    # cap the search window per event (trading days)

# ─── Entry-timing backtest config ───
ENTRY_OFFSETS    = [1, 2, 3, 5, 7, 10, 15, 20]  # trading days before the ex-date
MIN_ENTRY_EVENTS = 3            # need at least this many past events to call a sweet spot

# ─── Scan performance ───
MAX_WORKERS = 8                 # parallel requests to Yahoo Finance


def compute_rebound(hist: pd.DataFrame):
    """
    How long the stock historically takes to recover to its pre-ex-dividend
    price. Returns (avg_days, worst_display, success_rate, expected_days, measured).
      - avg_days       : average trading days to recover (recovered events only)
      - worst_display  : longest recovery; ">60" if any event never recovered
      - success_rate   : fraction of measured events that recovered in the window
      - expected_days  : average holding INCLUDING failures charged at the full
                         60-day cap. This drives the efficiency score, so a
                         stock that often fails to recover rotates slower.
    Returns (None, None, None, None, 0) if there isn't enough history.
    """
    if hist is None or hist.empty or "Dividends" not in hist.columns:
        return None, None, None, None, 0

    closes = hist["Close"]
    ex_dates = list(hist.index[hist["Dividends"] > 0])
    if not ex_dates:
        return None, None, None, None, 0

    ex_dates = ex_dates[-REBOUND_LOOKBACK_EVENTS:]
    rebound_days, failures, measured = [], 0, 0

    for ex_date in ex_dates:
        pre = closes[closes.index < ex_date]
        if pre.empty:
            continue
        pre_price = pre.iloc[-1]
        post = closes[closes.index >= ex_date].iloc[:REBOUND_MAX_DAYS + 1]
        if len(post) < 2:
            continue
        # Skip events too recent to judge fairly: if the window is still open
        # and the price hasn't recovered yet, we can't call it a failure.
        window_complete = len(post) >= REBOUND_MAX_DAYS + 1
        recovered = post[post >= pre_price]
        if recovered.empty:
            if not window_complete:
                continue  # jury still out - don't count it either way
            measured += 1
            failures += 1
            continue
        measured += 1
        days = post.index.get_loc(recovered.index[0])
        rebound_days.append(days)

    if measured == 0:
        return None, None, None, None, 0

    avg = (sum(rebound_days) / len(rebound_days)) if rebound_days else None
    if failures > 0:
        worst_display = f">{REBOUND_MAX_DAYS}"
    else:
        worst_display = max(rebound_days) if rebound_days else None
    success_rate = (measured - failures) / measured
    expected_days = (sum(rebound_days) + failures * REBOUND_MAX_DAYS) / measured
    return avg, worst_display, success_rate, expected_days, measured


def compute_entry_and_drop(hist: pd.DataFrame):
    """
    Back-tests entry timing and measures the real ex-date price drop.

    Entry timing: for each past ex-dividend event, simulates buying at the
    close 1, 2, 3, 5, 7, 10, 15 or 20 trading days before the ex-date and
    holding through it (valued at the ex-date close plus the dividend).
    The offset with the best average return is the stock's historical
    sweet spot for getting in.

    Drop ratio: (close the day before the ex-date - open on the ex-date)
    divided by the dividend. 1.0 means the price fell by exactly the
    dividend; below 1.0 means it gave up less than the dividend.

    Returns (best_entry_days, best_entry_avg_ret, drop_ratio) - any of them
    None if there isn't enough history to measure.
    """
    if hist is None or hist.empty or "Dividends" not in hist.columns:
        return None, None, None

    closes = hist["Close"]
    opens = hist["Open"] if "Open" in hist.columns else closes
    divs = hist["Dividends"]
    ex_positions = [i for i, d in enumerate(divs) if d > 0]
    if not ex_positions:
        return None, None, None
    ex_positions = ex_positions[-REBOUND_LOOKBACK_EVENTS:]

    entry_returns = {n: [] for n in ENTRY_OFFSETS}
    drop_ratios = []

    for pos in ex_positions:
        div = divs.iloc[pos]
        if pos < 1 or div <= 0:
            continue
        pre_close = closes.iloc[pos - 1]
        ex_open = opens.iloc[pos]
        if pre_close > 0 and ex_open > 0:
            drop_ratios.append((pre_close - ex_open) / div)
        for n in ENTRY_OFFSETS:
            if pos - n < 0:
                continue
            entry_price = closes.iloc[pos - n]
            if entry_price <= 0:
                continue
            ret = (closes.iloc[pos] + div - entry_price) / entry_price
            entry_returns[n].append(ret)

    best_n, best_ret = None, None
    for n in ENTRY_OFFSETS:
        rets = entry_returns[n]
        if len(rets) < MIN_ENTRY_EVENTS:
            continue
        avg = sum(rets) / len(rets)
        if best_ret is None or avg > best_ret:
            best_n, best_ret = n, avg

    drop_ratio = (sum(drop_ratios) / len(drop_ratios)) if drop_ratios else None
    return best_n, best_ret, drop_ratio


def compute_trend(hist: pd.DataFrame) -> str:
    """
    Momentum read from moving averages on dividend-adjusted (total-return)
    prices. Using adjusted prices matters: a 20%-yield fund's raw price drifts
    down by design even when holders are not losing money.
    """
    if hist is None or hist.empty:
        return "N/A"
    col = "Adj Close" if "Adj Close" in hist.columns else "Close"
    closes = hist[col].dropna()
    if len(closes) < 200:
        return "N/A"
    price = closes.iloc[-1]
    sma50 = closes.tail(50).mean()
    sma200 = closes.tail(200).mean()
    if price > sma200 and sma50 > sma200:
        return "Up"
    if price < sma200 and sma50 < sma200:
        return "Down"
    return "Mixed"


def liquidity_label(avg_volume: float) -> str:
    if avg_volume < THIN_VOLUME:
        return "Thin"
    if avg_volume >= HIGH_VOLUME:
        return "High"
    return "Moderate"


def fetch_candidate(ticker: str, today: datetime):
    """Returns (row_dict, note). Exactly one of the two is None; both None
    means the stock was simply filtered out. A note starting with 'EXCLUDED:'
    means the name was dropped for its sector rather than for a data error."""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        ex_date_ts = info.get("exDividendDate")
        if not ex_date_ts:
            return None, None
        # Yahoo timestamps are UTC; interpret them as UTC so the calendar
        # date doesn't shift depending on the machine's time zone.
        ex_date = datetime.fromtimestamp(ex_date_ts, tz=timezone.utc).replace(tzinfo=None)
        days_until = (ex_date.date() - today.date()).days
        if not (0 < days_until <= EX_DATE_WINDOW_DAYS):
            return None, None

        price = info.get("currentPrice") or info.get("regularMarketPrice")
        if not price or not (MIN_STOCK_PRICE <= price <= MAX_STOCK_PRICE):
            return None, None

        div_rate = info.get("dividendRate")   # annual dividend in $
        if not div_rate:
            return None, None
        annual_yield = div_rate / price
        if annual_yield < MIN_ANNUAL_YIELD:
            return None, None

        # Sector, with a fund/ETF fallback since Yahoo leaves those blank.
        sector = info.get("sector") or FUND_SECTORS.get(ticker) or "Unknown"
        if sector in EXCLUDED_SECTORS:
            return None, f"EXCLUDED:{ticker} ({sector})"

        avg_volume = info.get("averageVolume") or 0
        if avg_volume < MIN_AVG_VOLUME:
            return None, None

        # Expensive step: history, only for stocks that passed the cheap filters
        hist = stock.history(period="2y", auto_adjust=False)

        freq = 4
        if "Dividends" in hist.columns:
            one_year_ago = hist.index.max() - pd.Timedelta(days=365)
            recent = hist.index[(hist["Dividends"] > 0) & (hist.index >= one_year_ago)]
            if len(recent) > 0:
                freq = len(recent)
        div_per_capture = div_rate / freq
        capture_yield = div_per_capture / price

        avg_reb, worst_display, success, expected_days, events = compute_rebound(hist)
        best_entry, best_entry_ret, drop_ratio = compute_entry_and_drop(hist)
        trend = compute_trend(hist)
        liq = liquidity_label(avg_volume)

        # Round-lot economics, after tax at both rates
        capital = price * ROUND_LOT
        div_dollars = div_per_capture * ROUND_LOT
        aftertax_high = div_dollars * (1 - TAX_RATE_HIGH)
        aftertax_low = div_dollars * (1 - TAX_RATE_LOW)

        # Annualised efficiency (models continuous capital rotation).
        # Uses expected_days, which charges failed recoveries at the full
        # 60-day wait, so unreliable rebounds drag the number down honestly.
        # Taxed at the higher (conservative) rate.
        if expected_days is not None:
            holding = expected_days + BUY_BUFFER_DAYS
            per_cycle = (div_per_capture * (1 - TAX_RATE_HIGH)) / price
            cycles = TRADING_DAYS / max(holding, 1)
            ann_eff = per_cycle * cycles * 100
        else:
            ann_eff = None

        beta = info.get("beta")
        # Yahoo occasionally returns junk for a name (e.g. a bare number);
        # fall back to the long name, then the ticker itself.
        name = info.get("shortName") or info.get("longName") or ticker
        if str(name).replace(".", "").replace(",", "").isdigit():
            name = info.get("longName") or ticker

        row = {
            "Ticker":        ticker,
            "Name":          name,
            "Pref":          "Yes" if ticker in FAVORED_TICKERS else "",
            "Sector":        sector,
            "Price":         round(price, 2),
            "Ex-Date":       ex_date.strftime("%Y-%m-%d"),
            "Days Until":    days_until,
            "Annual Yield":  f"{annual_yield:.2%}",
            "Capture Yield": f"{capture_yield:.2%}",
            "Div/100sh":     round(div_dollars, 2),
            "Capital":       int(round(capital, 0)),
            COL_AT_HIGH:     round(aftertax_high, 2),
            COL_AT_LOW:      round(aftertax_low, 2),
            "Avg Reb":       round(avg_reb, 1) if avg_reb is not None else "N/A",
            "Worst Reb":     worst_display if worst_display is not None else "N/A",
            "Recovered":     f"{success:.0%}" if success is not None else "N/A",
            "Best Entry":    f"{best_entry}d" if best_entry is not None else "N/A",
            "Drop Ratio":    round(drop_ratio, 2) if drop_ratio is not None else "N/A",
            "Trend":         trend,
            "Liquidity":     liq,
            "Beta":          round(beta, 2) if beta is not None else "N/A",
            "Ann.Eff%":      round(ann_eff, 1) if ann_eff is not None else "N/A",
            # private fields for scoring and the rotation plan
            "_ann_eff":      ann_eff,
            "_trend":        trend,
            "_liq":          liq,
            "_sector":       sector,
            "_ex_date":      ex_date.date(),
            "_price":        price,
            "_div":          div_per_capture,
            "_best_entry":   best_entry,
            "_avg_reb":      avg_reb,
            "_worst":        worst_display,
        }
        return row, None
    except Exception as e:
        return None, f"{ticker}: {type(e).__name__}"


def score(row: dict) -> float:
    """
    Ranking score = annualised rotation efficiency, discounted for a
    downtrend (price may not recover) and for thin liquidity (hard to exit).
    """
    ann = row["_ann_eff"]
    if ann is None:
        return 0.0
    factor = 1.0
    if row["_trend"] == "Down":
        factor *= 0.6
    elif row["_trend"] == "Mixed":
        factor *= 0.85
    if row["_liq"] == "Thin":
        factor *= 0.85
    return ann * factor


def slice_states(today64):
    """
    Turns HOLDINGS into the starting state of each slice.

    Every slice is a pot of CAPITAL that holds one position at a time. A slice
    holding a position with a known sell date is busy until the cash settles;
    a slice holding a position with no sell date (stuck below its entry price)
    is BLOCKED and gets no trades scheduled at all, because pretending the
    money is available is what makes a plan unfundable.
    """
    states = {s: {"free": today64, "blocked": None, "holding": None, "hold_until": None}
              for s in range(1, N_SLICES + 1)}
    for h in HOLDINGS:
        s = h.get("slice")
        if s not in states:
            continue
        states[s]["holding"] = h["ticker"]
        if h.get("free"):
            free = np.datetime64(h["free"], "D")
            if free > states[s]["free"]:
                states[s]["free"] = free
            # kept separately: "free" moves as trades are scheduled, this does not
            states[s]["hold_until"] = states[s]["free"]
        else:
            states[s]["blocked"] = h["ticker"]
    return states


def build_rotation_plan(df: pd.DataFrame, today: datetime):
    """
    Builds a chronological trade calendar for N_SLICES independent pots of
    CAPITAL, the way the strategy is actually traded:

      buy on the stock's Best Entry day -> hold through the ex-date ->
      sell once the price is back to even (its average rebound) ->
      one settlement day -> straight into the next candidate.

    Each slice rotates on its own clock, so one position that gets stuck does
    not stall the others. Greedy choice: whenever a slice's capital is free,
    look at the candidates it could buy within the next 5 trading days and take
    the one with the best Score. No two slices are ever put in the same stock,
    and anything already in HOLDINGS is off the table.

    Dollar figures use whole 100-share lots bought with one slice of CAPITAL,
    minus any FINANCE_FEE on the gross dividend, then tax on what remains. Net
    dollars are shown at both tax rates side by side.

    Returns (plan_df, states) so the caller can report blocked slices.
    """
    today64 = np.datetime64(today.date(), "D")
    states = slice_states(today64)
    held = {h["ticker"] for h in HOLDINGS}

    entries = []
    for _, row in df.iterrows():
        if row["Ticker"] in held:
            continue  # already own it - the account does not double up
        if row["_avg_reb"] is None or row["Score"] <= 0:
            continue  # no reliable rebound history - can't schedule it honestly
        lots = int(CAPITAL // (row["_price"] * ROUND_LOT))
        if lots < 1:
            continue  # 100 shares cost more than one slice of capital
        shares = lots * ROUND_LOT
        ex = np.datetime64(row["_ex_date"], "D")
        entry_days = row["_best_entry"] if row["_best_entry"] is not None else BUY_BUFFER_DAYS
        buy = np.busday_offset(ex, -entry_days, roll="backward")
        if buy < today64:
            buy = np.busday_offset(today64, 0, roll="forward")  # can still buy: ex-date is ahead
        sell = np.busday_offset(ex, ceil(row["_avg_reb"]), roll="forward")
        free = np.busday_offset(sell, SETTLEMENT_DAYS, roll="forward")
        gross = shares * row["_div"]
        after_fee = gross * (1 - FINANCE_FEE)
        entries.append({
            "ticker": row["Ticker"], "buy": buy, "ex": ex, "sell": sell,
            "free": free, "shares": shares, "cost": shares * row["_price"],
            "gross": gross,
            "net_high": after_fee * (1 - TAX_RATE_HIGH),
            "net_low": after_fee * (1 - TAX_RATE_LOW),
            "score": row["Score"], "worst": row["_worst"], "sector": row["_sector"],
            "pref": row["Ticker"] in FAVORED_TICKERS,
        })

    def best_for(free_date):
        """The trade this slice would take next, or None."""
        feasible = []
        for e in entries:
            # Still buyable as long as the cash frees at least one trading day
            # before the ex-date. If the ideal Best Entry day has already
            # passed, buy as soon as the money is free instead.
            last_buy = np.busday_offset(e["ex"], -1, roll="backward")
            if last_buy < free_date:
                continue
            feasible.append((max(e["buy"], free_date), e))
        if not feasible:
            return None
        earliest = min(ab for ab, _ in feasible)
        window_end = np.busday_offset(earliest, 5, roll="forward")
        pool = [(ab, e) for ab, e in feasible if ab <= window_end]
        # Preferred names win close calls: their score gets a selection-only
        # edge (PREFERRED_EDGE). The displayed Score stays untouched.
        return max(pool, key=lambda p: p[1]["score"] * (1 + PREFERRED_EDGE if p[1]["pref"] else 0))

    booked = []
    while True:
        # Whichever free slice can act soonest goes first.
        best = None
        for s, st in sorted(states.items()):
            if st["blocked"] is not None:
                continue
            choice = best_for(st["free"])
            if choice is None:
                continue
            buy_date, pick = choice
            if best is None or buy_date < best[0]:
                best = (buy_date, s, pick)
        if best is None:
            break
        buy_date, s, pick = best
        booked.append({"slice": s, "buy": buy_date, "pick": pick})
        states[s]["free"] = pick["free"]
        entries = [e for e in entries if e["ticker"] != pick["ticker"]]

    booked.sort(key=lambda b: (b["buy"], b["slice"]))
    plan, running_high, running_low = [], 0.0, 0.0
    for b in booked:
        pick = b["pick"]
        running_high += pick["net_high"]
        running_low += pick["net_low"]
        plan.append({
            "Slice":      b["slice"],
            "Buy":        str(b["buy"]),
            "Ticker":     pick["ticker"],
            "Pref":       "Yes" if pick["pref"] else "",
            "Sector":     pick["sector"],
            "Ex-Date":    str(pick["ex"]),
            "Est. Sell":  str(pick["sell"]),
            "Cash Free":  str(pick["free"]),
            "Worst Reb":  pick["worst"] if pick["worst"] is not None else "N/A",
            "Shares":     pick["shares"],
            "Cost $":     int(round(pick["cost"])),
            "Gross Div $": round(pick["gross"], 2),
            COL_NET_HIGH: round(pick["net_high"], 2),
            COL_NET_LOW:  round(pick["net_low"], 2),
            COL_RUN_HIGH: round(running_high, 2),
            COL_RUN_LOW:  round(running_low, 2),
        })

    plan_df = pd.DataFrame(plan)
    if not plan_df.empty:
        plan_df.index += 1
    return plan_df, states


def sector_overlaps(plan_df: pd.DataFrame):
    """
    Finds dates where two or more slices would be holding the same sector at
    the same time. Four pots picking from a list dominated by rate-sensitive
    income names can quietly turn into one big bet on interest rates, which is
    the opposite of what running separate slices is for.

    Returns a list of (sector, [tickers]) for each overlapping sector.
    """
    if plan_df.empty:
        return []
    rows = []
    for _, r in plan_df.iterrows():
        rows.append((r["Sector"], r["Ticker"],
                     np.datetime64(r["Buy"], "D"), np.datetime64(r["Cash Free"], "D")))
    flagged = {}
    for i, (sec_a, tick_a, buy_a, free_a) in enumerate(rows):
        for sec_b, tick_b, buy_b, free_b in rows[i + 1:]:
            if sec_a != sec_b:
                continue
            if buy_a <= free_b and buy_b <= free_a:   # the holds overlap in time
                flagged.setdefault(sec_a, set()).update([tick_a, tick_b])
    return [(sec, sorted(ticks)) for sec, ticks in sorted(flagged.items())]


def save_calendar_html(plan_df: pd.DataFrame, today: datetime, out_file: str, states=None):
    """
    Visual month-grid calendar of the rotation plan, saved as a dated local
    HTML file (rotation_calendar_YYYYMMDD.html, matching the CSV and Excel
    outputs) that opens in any browser. Each slice gets its own colour; buy,
    ex-date, sell, and cash-free days are labelled, and the days in between
    are tinted so it is obvious when that slice's capital is occupied.
    """
    import calendar as cal
    from datetime import date as ddate, timedelta

    def parse(s):
        y, m, d = map(int, str(s).split("-"))
        return ddate(y, m, d)

    palette = ["#60a5fa", "#4ade80", "#fbbf24", "#a78bfa", "#fb7185"]
    chips, tints = {}, {}
    for _, row in plan_df.iterrows():
        sl = int(row["Slice"])
        color = palette[(sl - 1) % len(palette)]
        buy, ex = parse(row["Buy"]), parse(row["Ex-Date"])
        sell, free = parse(row["Est. Sell"]), parse(row["Cash Free"])
        t = f'{row["Ticker"]} S{sl}'
        chips.setdefault(buy, []).append((f"BUY {t}", color))
        chips.setdefault(ex, []).append((f"EX-DATE {t}", color))
        chips.setdefault(sell, []).append((f"SELL {t}", color))
        chips.setdefault(free, []).append((f"CASH FREE S{sl}", color))
        cur = buy
        while cur <= sell:
            if cur.weekday() < 5:
                tints.setdefault(cur, color)
            cur += timedelta(days=1)

    all_days = list(chips.keys())
    start, end = min(all_days), max(all_days)
    months = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append((y, m))
        m += 1
        if m == 13:
            y, m = y + 1, 1

    total_high = plan_df[COL_RUN_HIGH].iloc[-1]
    total_low = plan_df[COL_RUN_LOW].iloc[-1]
    total_capital = CAPITAL * N_SLICES
    parts = [f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>Rotation Plan Calendar</title><style>
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#0f1117;color:#e2e8f0;padding:30px 16px;}}
.wrap{{max-width:900px;margin:0 auto;}}
h1{{font-size:1.5rem;color:#f8fafc;}}
.sub{{color:#64748b;font-size:0.9rem;margin:6px 0 20px;}}
.legend{{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:24px;font-size:0.8rem;color:#94a3b8;}}
.dot{{display:inline-block;width:10px;height:10px;border-radius:50%;margin-right:5px;vertical-align:middle;}}
.note{{background:#1c1508;border-left:3px solid #fbbf24;color:#cbb994;padding:10px 14px;margin-bottom:22px;font-size:0.82rem;}}
table{{width:100%;border-collapse:collapse;margin-bottom:30px;table-layout:fixed;}}
caption{{text-align:left;font-size:1.05rem;font-weight:700;color:#f1f5f9;padding:8px 0;}}
th{{color:#64748b;font-size:0.7rem;text-transform:uppercase;padding:6px;border-bottom:1px solid #252d3d;}}
td{{border:1px solid #1c2333;vertical-align:top;height:64px;padding:4px;font-size:0.78rem;color:#64748b;}}
td .n{{font-weight:600;color:#94a3b8;}}
.chip{{display:block;margin-top:3px;padding:1px 5px;border-radius:4px;font-size:0.66rem;font-weight:700;color:#0f1117;}}
.we{{background:#12151e;}}
</style></head><body><div class="wrap">
<h1>Rotation Plan Calendar</h1>
<div class="sub">Generated {today.strftime('%Y-%m-%d')} &middot; {N_SLICES} slices of ${CAPITAL:,} (${total_capital:,} in total) &middot; estimated total net ${total_high:,.2f} at {TAX_RATE_HIGH:.0%} tax, ${total_low:,.2f} at {TAX_RATE_LOW:.0%} tax</div>
"""]

    blocked = [(s, st) for s, st in sorted((states or {}).items()) if st["blocked"]]
    busy = [(s, st) for s, st in sorted((states or {}).items())
            if st["blocked"] is None and st["holding"]]
    if blocked or busy:
        lines = []
        for s, st in blocked:
            lines.append(f'Slice {s} is blocked: it still holds {st["blocked"]}, '
                         f'which has no sell date yet. No trades are scheduled for it.')
        for s, st in busy:
            lines.append(f'Slice {s} holds {st["holding"]} until {str(st["hold_until"])}, '
                         f'so its first trade below starts from that date.')
        parts.append('<div class="note">' + '<br>'.join(lines) + '</div>')

    parts.append('<div class="legend">')
    for _, row in plan_df.iterrows():
        sl = int(row["Slice"])
        color = palette[(sl - 1) % len(palette)]
        star = "&#9733; " if row.get("Pref") == "Yes" else ""
        parts.append(f'<span><span class="dot" style="background:{color};"></span>'
                     f'{star}S{sl} {row["Ticker"]} (net ${row[COL_NET_HIGH]:,.2f} to ${row[COL_NET_LOW]:,.2f})</span>')
    parts.append('</div>')

    for (yy, mm) in months:
        parts.append(f'<table><caption>{cal.month_name[mm]} {yy}</caption><tr>'
                     + ''.join(f'<th>{w}</th>' for w in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]) + '</tr>')
        for week in cal.Calendar().monthdatescalendar(yy, mm):
            parts.append('<tr>')
            for day in week:
                if day.month != mm:
                    parts.append('<td class="we"></td>')
                    continue
                style = ""
                if day in tints:
                    style = f' style="background:{tints[day]}22;"'
                cls = ' class="we"' if day.weekday() >= 5 else ''
                cell = f'<td{cls}{style}><span class="n">{day.day}</span>'
                for label, color in chips.get(day, []):
                    cell += f'<span class="chip" style="background:{color};">{label}</span>'
                cell += '</td>'
                parts.append(cell)
            parts.append('</tr>')
        parts.append('</table>')

    parts.append('<div class="sub">S1 to S4 = which slice of capital pays for the trade. Each slice holds '
                 'one position at a time and rotates on its own clock. '
                 '&#9733; = a preferred name: judged safe to keep holding if the rebound runs long. '
                 'Tinted days = that slice\'s capital is occupied. '
                 'Sell days are estimates based on each stock\'s average rebound time. '
                 f'Net figures show a range: taxed at {TAX_RATE_HIGH:.0%} (conservative) to {TAX_RATE_LOW:.0%} (optimistic).</div>')
    parts.append('</div></body></html>')

    with open(out_file, "w", encoding="utf-8") as f:
        f.write(''.join(parts))


def save_excel(shown: pd.DataFrame, plan_df: pd.DataFrame, out_file: str):
    """Formatted Excel copy of the results: bold frozen header, sized columns,
    centred data (long names stay left-aligned). Two sheets: the ranked
    candidates and the rotation plan."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def style_sheet(ws, left_aligned=("Name",)):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        skip_cols = set()
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            if cell.value in left_aligned:
                skip_cols.add(cell.column)
        ws.freeze_panes = "A2"
        centre = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column not in skip_cols:
                    cell.alignment = centre
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 35)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        shown.to_excel(writer, sheet_name="Candidates", index=True, index_label="#")
        style_sheet(writer.sheets["Candidates"])
        if plan_df is not None and not plan_df.empty:
            plan_df.to_excel(writer, sheet_name="Rotation Plan", index=True, index_label="#")
            style_sheet(writer.sheets["Rotation Plan"])


def run_scanner():
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    total_capital = CAPITAL * N_SLICES
    print(f"\nDividend Capture Scanner v10 - {today.strftime('%Y-%m-%d')}")
    print(f"Looking for ex-dates within the next {EX_DATE_WINDOW_DAYS} days")
    print(f"After-tax figures shown at two rates: {TAX_RATE_HIGH:.0%} (conservative) and {TAX_RATE_LOW:.0%} (optimistic)")
    fee_note = f", financing fee {FINANCE_FEE:.0%}" if FINANCE_FEE > 0 else ""
    print(f"Rotation capital: {N_SLICES} slices of ${CAPITAL:,} = ${total_capital:,}{fee_note}")
    if HOLDINGS:
        print(f"Currently held ({len(HOLDINGS)}): "
              + ", ".join(f"{h['ticker']} in slice {h['slice']}" for h in HOLDINGS)
              + " - these are excluded from the plan")
    if EXCLUDED_SECTORS:
        print(f"Excluded sectors: {', '.join(EXCLUDED_SECTORS)}")
    print(f"Preferred list: {len(FAVORED_TICKERS)} names flagged and favored in the plan\n")

    print("Fetching universe (S&P 500 + high-yield names + CEFs + preferreds)...")
    tickers = get_universe()
    print(f"Scanning {len(tickers)} stocks with {MAX_WORKERS} parallel workers...")
    print("(Stocks that pass the basic filters also get a rebound, trend,")
    print(" entry-timing, and recovery analysis from 2 years of history.)\n")

    candidates, errors, excluded = [], [], []
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(fetch_candidate, t, today): t for t in tickers}
        for future in as_completed(futures):
            done += 1
            row, note = future.result()
            if row:
                candidates.append(row)
            elif note and note.startswith("EXCLUDED:"):
                excluded.append(note[len("EXCLUDED:"):])
            elif note:
                errors.append(note)
            if done % 100 == 0:
                print(f"  {done}/{len(tickers)} scanned, {len(candidates)} candidates so far...")

    if excluded:
        print(f"\nExcluded by sector ({len(excluded)}): {', '.join(sorted(excluded))}")

    if errors:
        print(f"\nNote: {len(errors)} tickers were skipped due to data errors.")
        if len(errors) <= 10:
            for e in errors:
                print(f"  - {e}")

    if not candidates:
        print("\nNo candidates found matching the current filters.")
        if len(errors) > len(tickers) * 0.5:
            print("More than half the tickers errored - Yahoo Finance may be "
                  "blocking requests or has changed its data format.")
        return

    df = pd.DataFrame(candidates)
    df["Score"] = df.apply(score, axis=1).round(1)
    df = df.sort_values("Score", ascending=False).reset_index(drop=True)
    df.index += 1

    display_cols = [
        "Ticker", "Name", "Pref", "Sector", "Price", "Ex-Date", "Days Until",
        "Annual Yield", "Capture Yield", "Div/100sh", "Capital",
        COL_AT_HIGH, COL_AT_LOW,
        "Avg Reb", "Worst Reb", "Recovered", "Best Entry", "Drop Ratio",
        "Trend", "Liquidity", "Beta", "Ann.Eff%", "Score",
    ]
    shown = df[display_cols]

    print(f"\n{'='*140}")
    print(f"  TOP DIVIDEND CAPTURE CANDIDATES  ({len(df)} found)")
    print(f"{'='*140}")
    print(shown.to_string())

    print(f"\nColumn guide:")
    print(f"  Pref          = on the preferred list: names judged safe to keep holding if a")
    print(f"                  rebound takes far longer than expected; they win close calls")
    print(f"                  in the Rotation Plan (see FAVORED_TICKERS at the top of the script)")
    print(f"  Sector        = the stock's sector. Whole sectors can be ruled out (see")
    print(f"                  EXCLUDED_SECTORS); the plan also warns when slices overlap in one")
    print(f"  Annual Yield  = full year's dividend as % of price (ranked by this, higher is better)")
    print(f"  Capture Yield = single dividend payment as % of price (what you collect per trade)")
    print(f"  Div/100sh     = dividend dollars for one 100-share round lot")
    print(f"  Capital       = cost of 100 shares")
    print(f"  {COL_AT_HIGH}= dividend dollars per 100 shares after {TAX_RATE_HIGH:.0%} tax (the conservative estimate)")
    print(f"  {COL_AT_LOW}= the same after {TAX_RATE_LOW:.0%} tax (the optimistic estimate)")
    print(f"  Avg/Worst Reb = average and worst-case trading days to recover to the pre-ex price")
    print(f"                  ('>{REBOUND_MAX_DAYS}' means at least one past event never recovered in the window)")
    print(f"  Recovered     = how often it recovered within {REBOUND_MAX_DAYS} trading days")
    print(f"  Best Entry    = back-tested sweet spot: how many trading days before the ex-date")
    print(f"                  buying has historically worked best for this stock")
    print(f"  Drop Ratio    = how much of the dividend the price actually gives up on the ex-date")
    print(f"                  (1.0 = drops the full dividend; below 1.0 = drops less, which is good)")
    print(f"  Trend         = total-return price trend (Up / Down / Mixed). Downtrends are discouraged.")
    print(f"  Liquidity     = Thin / Moderate / High (thin = harder to exit)")
    print(f"  Ann.Eff%      = theoretical annualised after-tax return with continuous rotation,")
    print(f"                  taxed at {TAX_RATE_HIGH:.0%}; failed recoveries are charged at the full {REBOUND_MAX_DAYS}-day wait")
    print(f"  Score         = Ann.Eff% adjusted down for downtrend and thin liquidity")

    plan_df, states = build_rotation_plan(df, today)

    blocked = [(s, st) for s, st in sorted(states.items()) if st["blocked"]]
    busy = [(s, st) for s, st in sorted(states.items())
            if st["blocked"] is None and st["holding"]]
    if blocked or busy:
        print(f"\n{'='*140}")
        print("  SLICE STATUS")
        print(f"{'='*140}")
        for s, st in blocked:
            print(f"  Slice {s}: BLOCKED - still holds {st['blocked']} with no sell date. "
                  f"No trades scheduled for this slice.")
        for s, st in busy:
            print(f"  Slice {s}: holds {st['holding']} until {str(st['hold_until'])}. "
                  f"Its first trade below starts from that date.")
        free_now = N_SLICES - len(blocked) - len(busy)
        print(f"  {free_now} of {N_SLICES} slices are free today "
              f"(${free_now * CAPITAL:,} of ${total_capital:,}).")

    if not plan_df.empty:
        total_high = plan_df[COL_RUN_HIGH].iloc[-1]
        total_low = plan_df[COL_RUN_LOW].iloc[-1]
        tradeable = (N_SLICES - len(blocked)) * CAPITAL
        print(f"\n{'='*140}")
        print(f"  ROTATION PLAN  ({N_SLICES} slices of ${CAPITAL:,}, net shown at "
              f"{TAX_RATE_HIGH:.0%} and {TAX_RATE_LOW:.0%} tax"
              f"{', ' + format(FINANCE_FEE, '.0%') + ' financing fee' if FINANCE_FEE > 0 else ''})")
        print(f"{'='*140}")
        print(plan_df.to_string())
        print(f"\n  {len(plan_df)} trades planned over the next {EX_DATE_WINDOW_DAYS} days "
              f"across {plan_df['Slice'].nunique()} slices.")
        print(f"  Estimated total net: ${total_high:,.2f} at {TAX_RATE_HIGH:.0%} tax "
              f"({total_high / tradeable:.2%} of the ${tradeable:,} actually tradeable) to "
              f"${total_low:,.2f} at {TAX_RATE_LOW:.0%} tax ({total_low / tradeable:.2%}).")
        print(f"  Slice = which pot pays for the trade. Buy = the stock's own back-tested Best")
        print(f"  Entry day. Est. Sell = ex-date plus its average rebound. Cash Free = one")
        print(f"  settlement day after the sale (T+{SETTLEMENT_DAYS}).")
        print(f"  Worst Reb shows the honest risk: the longest that stock has ever taken to recover.")

        overlaps = sector_overlaps(plan_df)
        if overlaps:
            print(f"\n  SECTOR CONCENTRATION WARNING")
            for sec, ticks in overlaps:
                print(f"    {sec}: {', '.join(ticks)} would be held at the same time.")
            print(f"    Separate slices are meant to spread risk. Holding one sector in several")
            print(f"    of them at once turns the whole allocation back into a single bet.")

        cal_file = f"rotation_calendar_{today.strftime('%Y%m%d')}.html"
        try:
            save_calendar_html(plan_df, today, cal_file, states)
            print(f"\nSaved to {cal_file} (visual calendar - open it in any browser)")
        except PermissionError:
            print(f"\nCould not save {cal_file} - close it in the browser/editor and rerun.")
    else:
        print("\nNo rotation plan could be built. Either no candidates have reliable")
        print("rebound history, or every slice is already committed.")

    out_file = f"candidates_{today.strftime('%Y%m%d')}.csv"
    shown.to_csv(out_file, index=True)
    print(f"\nSaved to {out_file}")

    out_xlsx = f"candidates_{today.strftime('%Y%m%d')}.xlsx"
    try:
        save_excel(shown, plan_df, out_xlsx)
        print(f"Saved to {out_xlsx} (formatted for Excel, includes the Rotation Plan sheet)")
    except PermissionError:
        print(f"Could not save {out_xlsx} - close the file if it is open in Excel and rerun.")
    except ImportError:
        print(f"Skipped {out_xlsx} - install openpyxl first:  pip install openpyxl")


if __name__ == "__main__":
    run_scanner()
