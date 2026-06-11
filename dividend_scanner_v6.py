"""
Dividend Capture Scanner  (v6)
Pulls upcoming ex-dividend dates from Yahoo Finance and ranks the best
dividend-capture candidates.

What it looks at:
  - Dividend yield        : ranked directly by percentage (higher is better).
  - Rebound time          : average AND worst-case trading days for the price
                            to climb back to its pre-ex-dividend level. Events
                            that never recovered show as ">60" and are counted
                            against the stock, not hidden.
  - Recovery rate         : how often it recovered within the test window.
  - Best entry timing     : back-tests buying 1 to 20 trading days before the
                            ex-date to find each stock's historical sweet spot.
  - Drop ratio            : how much the price actually falls on the ex-date
                            compared with the dividend paid. Below 1.0 means
                            the stock gives up less than the dividend.
  - Price trend           : up / down / mixed, from moving averages on
                            dividend-adjusted (total-return) prices, so big
                            payers aren't unfairly flagged as falling.
  - Liquidity             : flags thinly-traded names that are harder to exit.
  - Round-lot economics   : dividend dollars and capital required per 100 shares.
  - After-tax gain        : dividend-capture trades are short-term, taxed as
                            ordinary income (see TAX_RATE below).
  - Annualised efficiency : models continuous capital rotation - hold for the
                            rebound, sell at even-or-better, redeploy into the
                            next dividend. Failed recoveries are charged at the
                            full 60-day wait, so unreliable stocks score lower.
  - Rotation plan         : a chronological trade calendar showing how a fixed
                            pot of capital (see CAPITAL) can be rolled from one
                            dividend into the next, with estimated net dollars
                            per trade after tax and any financing fee.
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from math import ceil
import warnings
warnings.filterwarnings("ignore")

# ─── S&P 500 tickers (fetched from Wikipedia) ───
def get_sp500_tickers():
    import requests
    from io import StringIO
    url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; dividend-scanner/1.0)"}
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()
    tables = pd.read_html(StringIO(resp.text))
    symbols = tables[0]["Symbol"].tolist()
    # Yahoo uses dashes where Wikipedia uses dots (BRK.B -> BRK-B)
    return [s.replace(".", "-") for s in symbols]

# ─── High-yield income names (mostly outside the S&P 500) ───
HIGH_YIELD_TICKERS = [
    # Mortgage REITs (often 9-15% yield)
    "AGNC", "NLY", "STWD", "ABR", "RITM", "DX", "CIM", "TWO", "PMT", "ARI",
    "EFC", "MFA", "IVR", "ORC",
    # Business Development Companies (BDCs, often 8-12%)
    "ARCC", "MAIN", "FSK", "OBDC", "BXSL", "PSEC", "HTGC", "GBDC", "TSLX",
    "CSWC", "GAIN", "PFLT", "BBDC", "TCPC",
    # REITs (5-9% yield)
    "O", "WPC", "STAG", "NNN", "EPR", "GLPI", "OHI", "LTC", "SBRA", "ADC",
    "IRM", "VICI", "KRG", "GNL",
    # Energy infrastructure (6-8%)
    "KMI", "OKE", "WES", "ET", "EPD", "MPLX", "ENB",
    # Other high-yield blue chips
    "BTI", "VZ", "T", "MO", "PFE",
]

# Closed-end funds (CEFs) - commonly high yield (7-11%).
# Note: CLM and CRF were removed; their headline yields are largely a return
# of the investor's own capital, which makes them misleading for capture.
CEF_TICKERS = [
    "PDI", "PTY", "PDO", "UTF", "UTG", "USA", "ETV", "ETY", "BST", "DSL",
    "RA", "GOF", "ECC", "OXLC", "EVT", "BME", "HQH", "GAB",
    "PCN", "PFL", "BCAT",
]

# Preferred-stock exposure via clean ETF tickers (individual preferred symbols
# are unreliable on Yahoo; these are a dependable proxy for the asset class).
PREFERRED_TICKERS = ["PFF", "PFFA", "PFFD", "PGX", "PGF", "FPE", "PFXF"]

def get_universe():
    """S&P 500 + high-yield names + CEFs + preferreds, de-duplicated."""
    sp500 = get_sp500_tickers()
    combined = sp500 + HIGH_YIELD_TICKERS + CEF_TICKERS + PREFERRED_TICKERS
    return list(dict.fromkeys(combined))  # preserves order, dedupes

# ─── Filter config (adjust to taste) ───
MIN_ANNUAL_YIELD     = 0.03      # 3% floor
MIN_AVG_VOLUME       = 100_000   # lowered so CEFs/preferreds appear; thin ones get flagged
MIN_STOCK_PRICE      = 5.0       # dollars
MAX_STOCK_PRICE      = 500.0     # dollars
EX_DATE_WINDOW_DAYS  = 30        # how far ahead to look (wide enough to plan rotations)

# ─── Economics / tax config ───
TAX_RATE        = 0.40   # Short-term / ordinary-income rate. Dividend-capture
                         # trades are too short to qualify for the lower 15-20%
                         # long-term capital-gains rate. Adjust to your bracket.
ROUND_LOT       = 100    # shares - round lots are far easier to sell than odd lots
BUY_BUFFER_DAYS = 2      # days capital is tied up before the ex-date (rotation calc)
TRADING_DAYS    = 252    # trading days per year

# ─── Rotation-plan config ───
CAPITAL         = 50_000  # dollars available to rotate from trade to trade
FINANCE_FEE     = 0.00    # cut taken by a trade-financing company, if any
                          # (e.g. 0.10 = they keep 10% of the gross dividend)
SETTLEMENT_DAYS = 1       # trading days for sale proceeds to settle (T+1)

# ─── Liquidity rating thresholds (avg shares/day) ───
THIN_VOLUME = 500_000
HIGH_VOLUME = 2_000_000

# ─── Rebound-metric config ───
REBOUND_LOOKBACK_EVENTS = 8     # measure the last N ex-dividend events
REBOUND_MAX_DAYS        = 60    # cap the search window per event (trading days)

# ─── Entry-timing backtest config ───
ENTRY_OFFSETS    = [1, 2, 3, 5, 7, 10, 15, 20]  # trading days before the ex-date
MIN_ENTRY_EVENTS = 3            # need at least this many past events to call a sweet spot

# ─── Scan performance ───
MAX_WORKERS = 8                 # parallel requests to Yahoo Finance


def compute_rebound(hist: pd.DataFrame):
    """
    How long the stock historically takes to recover to its pre-ex-dividend
    price. Returns (avg_days, worst_display, success_rate, expected_days, measured).
      - avg_days       : average trading days to recover (recovered events only)
      - worst_display  : longest recovery; ">60" if any event never recovered
      - success_rate   : fraction of measured events that recovered in the window
      - expected_days  : average holding INCLUDING failures charged at the full
                         60-day cap. This drives the efficiency score, so a
                         stock that often fails to recover rotates slower.
    Returns (None, None, None, None, 0) if there isn't enough history.
    """
    if hist is None or hist.empty or "Dividends" not in hist.columns:
        return None, None, None, None, 0

    closes = hist["Close"]
    ex_dates = list(hist.index[hist["Dividends"] > 0])
    if not ex_dates:
        return None, None, None, None, 0

    ex_dates = ex_dates[-REBOUND_LOOKBACK_EVENTS:]
    rebound_days, failures, measured = [], 0, 0

    for ex_date in ex_dates:
        pre = closes[closes.index < ex_date]
        if pre.empty:
            continue
        pre_price = pre.iloc[-1]
        post = closes[closes.index >= ex_date].iloc[:REBOUND_MAX_DAYS + 1]
        if len(post) < 2:
            continue
        # Skip events too recent to judge fairly: if the window is still open
        # and the price hasn't recovered yet, we can't call it a failure.
        window_complete = len(post) >= REBOUND_MAX_DAYS + 1
        recovered = post[post >= pre_price]
        if recovered.empty:
            if not window_complete:
                continue  # jury still out - don't count it either way
            measured += 1
            failures += 1
            continue
        measured += 1
        days = post.index.get_loc(recovered.index[0])
        rebound_days.append(days)

    if measured == 0:
        return None, None, None, None, 0

    avg = (sum(rebound_days) / len(rebound_days)) if rebound_days else None
    if failures > 0:
        worst_display = f">{REBOUND_MAX_DAYS}"
    else:
        worst_display = max(rebound_days) if rebound_days else None
    success_rate = (measured - failures) / measured
    expected_days = (sum(rebound_days) + failures * REBOUND_MAX_DAYS) / measured
    return avg, worst_display, success_rate, expected_days, measured


def compute_entry_and_drop(hist: pd.DataFrame):
    """
    Back-tests entry timing and measures the real ex-date price drop.

    Entry timing: for each past ex-dividend event, simulates buying at the
    close 1, 2, 3, 5, 7, 10, 15 or 20 trading days before the ex-date and
    holding through it (valued at the ex-date close plus the dividend).
    The offset with the best average return is the stock's historical
    sweet spot for getting in.

    Drop ratio: (close the day before the ex-date - open on the ex-date)
    divided by the dividend. 1.0 means the price fell by exactly the
    dividend; below 1.0 means it gave up less than the dividend.

    Returns (best_entry_days, best_entry_avg_ret, drop_ratio) - any of them
    None if there isn't enough history to measure.
    """
    if hist is None or hist.empty or "Dividends" not in hist.columns:
        return None, None, None

    closes = hist["Close"]
    opens = hist["Open"] if "Open" in hist.columns else closes
    divs = hist["Dividends"]
    ex_positions = [i for i, d in enumerate(divs) if d > 0]
    if not ex_positions:
        return None, None, None
    ex_positions = ex_positions[-REBOUND_LOOKBACK_EVENTS:]

    entry_returns = {n: [] for n in ENTRY_OFFSETS}
    drop_ratios = []

    for pos in ex_positions:
        div = divs.iloc[pos]
        if pos < 1 or div <= 0:
            continue
        pre_close = closes.iloc[pos - 1]
        ex_open = opens.iloc[pos]
        if pre_close > 0 and ex_open > 0:
            drop_ratios.append((pre_close - ex_open) / div)
        for n in ENTRY_OFFSETS:
            if pos - n < 0:
                continue
            entry_price = closes.iloc[pos - n]
            if entry_price <= 0:
                continue
            ret = (closes.iloc[pos] + div - entry_price) / entry_price
            entry_returns[n].append(ret)

    best_n, best_ret = None, None
    for n in ENTRY_OFFSETS:
        rets = entry_returns[n]
        if len(rets) < MIN_ENTRY_EVENTS:
            continue
        avg = sum(rets) / len(rets)
        if best_ret is None or avg > best_ret:
            best_n, best_ret = n, avg

    drop_ratio = (sum(drop_ratios) / len(drop_ratios)) if drop_ratios else None
    return best_n, best_ret, drop_ratio


def compute_trend(hist: pd.DataFrame) -> str:
    """
    Momentum read from moving averages on dividend-adjusted (total-return)
    prices. Using adjusted prices matters: a 20%-yield fund's raw price drifts
    down by design even when holders are not losing money.
    """
    if hist is None or hist.empty:
        return "N/A"
    col = "Adj Close" if "Adj Close" in hist.columns else "Close"
    closes = hist[col].dropna()
    if len(closes) < 200:
        return "N/A"
    price = closes.iloc[-1]
    sma50 = closes.tail(50).mean()
    sma200 = closes.tail(200).mean()
    if price > sma200 and sma50 > sma200:
        return "Up"
    if price < sma200 and sma50 < sma200:
        return "Down"
    return "Mixed"


def liquidity_label(avg_volume: float) -> str:
    if avg_volume < THIN_VOLUME:
        return "Thin"
    if avg_volume >= HIGH_VOLUME:
        return "High"
    return "Moderate"


def fetch_candidate(ticker: str, today: datetime):
    """Returns (row_dict, error_str). Exactly one of the two is None;
    both None means the stock was simply filtered out."""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        ex_date_ts = info.get("exDividendDate")
        if not ex_date_ts:
            return None, None
        # Yahoo timestamps are UTC; interpret them as UTC so the calendar
        # date doesn't shift depending on the machine's time zone.
        ex_date = datetime.fromtimestamp(ex_date_ts, tz=timezone.utc).replace(tzinfo=None)
        days_until = (ex_date.date() - today.date()).days
        if not (0 < days_until <= EX_DATE_WINDOW_DAYS):
            return None, None

        price = info.get("currentPrice") or info.get("regularMarketPrice")
        if not price or not (MIN_STOCK_PRICE <= price <= MAX_STOCK_PRICE):
            return None, None

        div_rate = info.get("dividendRate")   # annual dividend in $
        if not div_rate:
            return None, None
        annual_yield = div_rate / price
        if annual_yield < MIN_ANNUAL_YIELD:
            return None, None

        avg_volume = info.get("averageVolume") or 0
        if avg_volume < MIN_AVG_VOLUME:
            return None, None

        # Expensive step: history, only for stocks that passed the cheap filters
        hist = stock.history(period="2y", auto_adjust=False)

        freq = 4
        if "Dividends" in hist.columns:
            one_year_ago = hist.index.max() - pd.Timedelta(days=365)
            recent = hist.index[(hist["Dividends"] > 0) & (hist.index >= one_year_ago)]
            if len(recent) > 0:
                freq = len(recent)
        div_per_capture = div_rate / freq
        capture_yield = div_per_capture / price

        avg_reb, worst_display, success, expected_days, events = compute_rebound(hist)
        best_entry, best_entry_ret, drop_ratio = compute_entry_and_drop(hist)
        trend = compute_trend(hist)
        liq = liquidity_label(avg_volume)

        # Round-lot economics
        capital = price * ROUND_LOT
        div_dollars = div_per_capture * ROUND_LOT
        aftertax_dollars = div_dollars * (1 - TAX_RATE)

        # Annualised efficiency (models continuous capital rotation).
        # Uses expected_days, which charges failed recoveries at the full
        # 60-day wait, so unreliable rebounds drag the number down honestly.
        if expected_days is not None:
            holding = expected_days + BUY_BUFFER_DAYS
            per_cycle = (div_per_capture * (1 - TAX_RATE)) / price
            cycles = TRADING_DAYS / max(holding, 1)
            ann_eff = per_cycle * cycles * 100
        else:
            ann_eff = None

        beta = info.get("beta")
        # Yahoo occasionally returns junk for a name (e.g. a bare number);
        # fall back to the long name, then the ticker itself.
        name = info.get("shortName") or info.get("longName") or ticker
        if str(name).replace(".", "").replace(",", "").isdigit():
            name = info.get("longName") or ticker

        row = {
            "Ticker":        ticker,
            "Name":          name,
            "Price":         round(price, 2),
            "Ex-Date":       ex_date.strftime("%Y-%m-%d"),
            "Days Until":    days_until,
            "Annual Yield":  f"{annual_yield:.2%}",
            "Capture Yield": f"{capture_yield:.2%}",
            "Div/100sh":     round(div_dollars, 2),
            "Capital":       int(round(capital, 0)),
            "AfterTax/100":  round(aftertax_dollars, 2),
            "Avg Reb":       round(avg_reb, 1) if avg_reb is not None else "N/A",
            "Worst Reb":     worst_display if worst_display is not None else "N/A",
            "Recovered":     f"{success:.0%}" if success is not None else "N/A",
            "Best Entry":    f"{best_entry}d" if best_entry is not None else "N/A",
            "Drop Ratio":    round(drop_ratio, 2) if drop_ratio is not None else "N/A",
            "Trend":         trend,
            "Liquidity":     liq,
            "Beta":          round(beta, 2) if beta is not None else "N/A",
            "Ann.Eff%":      round(ann_eff, 1) if ann_eff is not None else "N/A",
            # private fields for scoring and the rotation plan
            "_ann_eff":      ann_eff,
            "_trend":        trend,
            "_liq":          liq,
            "_ex_date":      ex_date.date(),
            "_price":        price,
            "_div":          div_per_capture,
            "_best_entry":   best_entry,
            "_avg_reb":      avg_reb,
            "_worst":        worst_display,
        }
        return row, None
    except Exception as e:
        return None, f"{ticker}: {type(e).__name__}"


def score(row: dict) -> float:
    """
    Ranking score = annualised rotation efficiency, discounted for a
    downtrend (price may not recover) and for thin liquidity (hard to exit).
    """
    ann = row["_ann_eff"]
    if ann is None:
        return 0.0
    factor = 1.0
    if row["_trend"] == "Down":
        factor *= 0.6
    elif row["_trend"] == "Mixed":
        factor *= 0.85
    if row["_liq"] == "Thin":
        factor *= 0.85
    return ann * factor


def build_rotation_plan(df: pd.DataFrame, today: datetime) -> pd.DataFrame:
    """
    Builds a chronological trade calendar that rolls one pot of CAPITAL from
    dividend to dividend, the way the strategy is actually traded:

      buy on the stock's Best Entry day -> hold through the ex-date ->
      sell once the price is back to even (its average rebound) ->
      one settlement day -> straight into the next candidate.

    Greedy choice: whenever the capital is free, look at the candidates that
    can be bought within the next 5 trading days and take the one with the
    best Score, so the money never sits idle waiting for a distant trade.

    Dollar figures use whole 100-share lots bought with CAPITAL, minus any
    FINANCE_FEE on the gross dividend, then TAX_RATE on what remains.
    """
    entries = []
    for _, row in df.iterrows():
        if row["_avg_reb"] is None or row["Score"] <= 0:
            continue  # no reliable rebound history - can't schedule it honestly
        lots = int(CAPITAL // (row["_price"] * ROUND_LOT))
        if lots < 1:
            continue  # 100 shares cost more than the available capital
        shares = lots * ROUND_LOT
        ex = np.datetime64(row["_ex_date"], "D")
        entry_days = row["_best_entry"] if row["_best_entry"] is not None else BUY_BUFFER_DAYS
        buy = np.busday_offset(ex, -entry_days, roll="backward")
        today64 = np.datetime64(today.date(), "D")
        if buy < today64:
            buy = np.busday_offset(today64, 0, roll="forward")  # can still buy: ex-date is ahead
        sell = np.busday_offset(ex, ceil(row["_avg_reb"]), roll="forward")
        free = np.busday_offset(sell, SETTLEMENT_DAYS, roll="forward")
        gross = shares * row["_div"]
        net = gross * (1 - FINANCE_FEE) * (1 - TAX_RATE)
        entries.append({
            "ticker": row["Ticker"], "buy": buy, "ex": ex, "sell": sell,
            "free": free, "shares": shares, "cost": shares * row["_price"],
            "gross": gross, "net": net, "score": row["Score"],
            "worst": row["_worst"],
        })

    plan, running = [], 0.0
    free_date = np.datetime64(today.date(), "D")
    while True:
        # A candidate is still buyable as long as the capital frees up at
        # least one trading day before its ex-date. If its ideal Best Entry
        # day has already passed, buy as soon as the cash is free instead.
        feasible = []
        for e in entries:
            last_buy = np.busday_offset(e["ex"], -1, roll="backward")
            if last_buy < free_date:
                continue
            actual_buy = max(e["buy"], free_date)
            feasible.append((actual_buy, e))
        if not feasible:
            break
        earliest = min(ab for ab, _ in feasible)
        window_end = np.busday_offset(earliest, 5, roll="forward")
        pool = [(ab, e) for ab, e in feasible if ab <= window_end]
        actual_buy, pick = max(pool, key=lambda p: p[1]["score"])
        running += pick["net"]
        plan.append({
            "Buy":        str(actual_buy),
            "Ticker":     pick["ticker"],
            "Ex-Date":    str(pick["ex"]),
            "Est. Sell":  str(pick["sell"]),
            "Cash Free":  str(pick["free"]),
            "Worst Reb":  pick["worst"] if pick["worst"] is not None else "N/A",
            "Shares":     pick["shares"],
            "Cost $":     int(round(pick["cost"])),
            "Gross Div $": round(pick["gross"], 2),
            "Net $":      round(pick["net"], 2),
            "Running Net $": round(running, 2),
        })
        free_date = pick["free"]
        entries = [e for e in entries if e["ticker"] != pick["ticker"]]

    plan_df = pd.DataFrame(plan)
    if not plan_df.empty:
        plan_df.index += 1
    return plan_df


def save_excel(shown: pd.DataFrame, plan_df: pd.DataFrame, out_file: str):
    """Formatted Excel copy of the results: bold frozen header, sized columns,
    centred data (long names stay left-aligned). Two sheets: the ranked
    candidates and the rotation plan."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def style_sheet(ws, left_aligned=("Name",)):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        skip_cols = set()
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            if cell.value in left_aligned:
                skip_cols.add(cell.column)
        ws.freeze_panes = "A2"
        centre = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column not in skip_cols:
                    cell.alignment = centre
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 35)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        shown.to_excel(writer, sheet_name="Candidates", index=True, index_label="#")
        style_sheet(writer.sheets["Candidates"])
        if plan_df is not None and not plan_df.empty:
            plan_df.to_excel(writer, sheet_name="Rotation Plan", index=True, index_label="#")
            style_sheet(writer.sheets["Rotation Plan"])


def run_scanner():
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    print(f"\nDividend Capture Scanner v6 - {today.strftime('%Y-%m-%d')}")
    print(f"Looking for ex-dates within the next {EX_DATE_WINDOW_DAYS} days")
    print(f"Tax rate assumed: {TAX_RATE:.0%} (short-term / ordinary income)")
    fee_note = f", financing fee {FINANCE_FEE:.0%}" if FINANCE_FEE > 0 else ""
    print(f"Rotation capital: ${CAPITAL:,}{fee_note}\n")

    print("Fetching universe (S&P 500 + high-yield names + CEFs + preferreds)...")
    tickers = get_universe()
    print(f"Scanning {len(tickers)} stocks with {MAX_WORKERS} parallel workers...")
    print("(Stocks that pass the basic filters also get a rebound, trend,")
    print(" entry-timing, and recovery analysis from 2 years of history.)\n")

    candidates, errors = [], []
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(fetch_candidate, t, today): t for t in tickers}
        for future in as_completed(futures):
            done += 1
            row, err = future.result()
            if row:
                candidates.append(row)
            elif err:
                errors.append(err)
            if done % 100 == 0:
                print(f"  {done}/{len(tickers)} scanned, {len(candidates)} candidates so far...")

    if errors:
        print(f"\nNote: {len(errors)} tickers were skipped due to data errors.")
        if len(errors) <= 10:
            for e in errors:
                print(f"  - {e}")

    if not candidates:
        print("\nNo candidates found matching the current filters.")
        if len(errors) > len(tickers) * 0.5:
            print("More than half the tickers errored - Yahoo Finance may be "
                  "blocking requests or has changed its data format.")
        return

    df = pd.DataFrame(candidates)
    df["Score"] = df.apply(score, axis=1).round(1)
    df = df.sort_values("Score", ascending=False).reset_index(drop=True)
    df.index += 1

    display_cols = [
        "Ticker", "Name", "Price", "Ex-Date", "Days Until",
        "Annual Yield", "Capture Yield", "Div/100sh", "Capital", "AfterTax/100",
        "Avg Reb", "Worst Reb", "Recovered", "Best Entry", "Drop Ratio",
        "Trend", "Liquidity", "Beta", "Ann.Eff%", "Score",
    ]
    shown = df[display_cols]

    print(f"\n{'='*130}")
    print(f"  TOP DIVIDEND CAPTURE CANDIDATES  ({len(df)} found)")
    print(f"{'='*130}")
    print(shown.to_string())

    print(f"\nColumn guide:")
    print(f"  Annual Yield  = full year's dividend as % of price (ranked by this, higher is better)")
    print(f"  Capture Yield = single dividend payment as % of price (what you collect per trade)")
    print(f"  Div/100sh     = dividend dollars for one 100-share round lot")
    print(f"  Capital       = cost of 100 shares")
    print(f"  AfterTax/100  = dividend dollars per 100 shares after {TAX_RATE:.0%} tax")
    print(f"  Avg/Worst Reb = average and worst-case trading days to recover to the pre-ex price")
    print(f"                  ('>{REBOUND_MAX_DAYS}' means at least one past event never recovered in the window)")
    print(f"  Recovered     = how often it recovered within {REBOUND_MAX_DAYS} trading days")
    print(f"  Best Entry    = back-tested sweet spot: how many trading days before the ex-date")
    print(f"                  buying has historically worked best for this stock")
    print(f"  Drop Ratio    = how much of the dividend the price actually gives up on the ex-date")
    print(f"                  (1.0 = drops the full dividend; below 1.0 = drops less, which is good)")
    print(f"  Trend         = total-return price trend (Up / Down / Mixed). Downtrends are discouraged.")
    print(f"  Liquidity     = Thin / Moderate / High (thin = harder to exit)")
    print(f"  Ann.Eff%      = theoretical annualised after-tax return with continuous rotation;")
    print(f"                  failed recoveries are charged at the full {REBOUND_MAX_DAYS}-day wait")
    print(f"  Score         = Ann.Eff% adjusted down for downtrend and thin liquidity")

    plan_df = build_rotation_plan(df, today)
    if not plan_df.empty:
        total_net = plan_df["Net $"].iloc[-1] if "Running Net $" not in plan_df else plan_df["Running Net $"].iloc[-1]
        print(f"\n{'='*130}")
        print(f"  ROTATION PLAN  (${CAPITAL:,} rolled from trade to trade, "
              f"{TAX_RATE:.0%} tax{', ' + format(FINANCE_FEE, '.0%') + ' financing fee' if FINANCE_FEE > 0 else ''})")
        print(f"{'='*130}")
        print(plan_df.to_string())
        print(f"\n  {len(plan_df)} trades planned over the next {EX_DATE_WINDOW_DAYS} days. "
              f"Estimated total net: ${total_net:,.2f} "
              f"({total_net / CAPITAL:.2%} of the capital).")
        print(f"  Buy = the stock's own back-tested Best Entry day. Est. Sell = ex-date plus its")
        print(f"  average rebound. Cash Free = one settlement day after the sale (T+{SETTLEMENT_DAYS}).")
        print(f"  Worst Reb shows the honest risk: the longest that stock has ever taken to recover.")
    else:
        print("\nNo rotation plan could be built (no candidates with reliable rebound history).")

    out_file = f"candidates_{today.strftime('%Y%m%d')}.csv"
    shown.to_csv(out_file, index=True)
    print(f"\nSaved to {out_file}")

    out_xlsx = f"candidates_{today.strftime('%Y%m%d')}.xlsx"
    try:
        save_excel(shown, plan_df, out_xlsx)
        print(f"Saved to {out_xlsx} (formatted for Excel, includes the Rotation Plan sheet)")
    except PermissionError:
        print(f"Could not save {out_xlsx} - close the file if it is open in Excel and rerun.")
    except ImportError:
        print(f"Skipped {out_xlsx} - install openpyxl first:  pip install openpyxl")


if __name__ == "__main__":
    run_scanner()
